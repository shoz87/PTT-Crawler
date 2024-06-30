import os
import requests
import threading
import time
import tkinter as tk
import sys
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from pathlib import Path
from bs4 import BeautifulSoup

if getattr(sys, 'frozen', False):
    os.chdir(sys._MEIPASS)

def get_articles_from_page(url, headers, keyword, session):
    retry_count = 0
    current_year = datetime.now().year  # 获取当前年份
    while retry_count < 5:
        try:
            response = session.get(url, headers=headers)
            if response.status_code == 200:
                break
            else:
                print(f"Failed to fetch page: {url} with status code: {response.status_code}")
                retry_count += 1
                time.sleep(2)
        except requests.exceptions.RequestException as e:
            print(f"Exception occurred while fetching page: {url} - {e}")
            retry_count += 1
            time.sleep(2)
    else:
        return [], None

    soup = BeautifulSoup(response.text, "html.parser")
    articles = soup.find_all("div", class_="r-ent")
    if not articles:
        print("No articles found on the page")

    data_list = []

    for article in articles:
        data = {}
        title_tag = article.find("div", class_="title")
        if title_tag and title_tag.a:
            title = title_tag.a.text.strip()
            link = "https://www.ptt.cc" + title_tag.a['href']
            title_link = f"{title} ({link})"
        else:
            title = "沒有標題"
            link = "N/A"
            title_link = title

        if keyword.lower() in title.lower():
            nrec = article.find("div", class_="nrec")
            if nrec and nrec.span:
                popularity = nrec.span.text.strip()
            else:
                popularity = "N/A"

            data["人氣"] = popularity
            date_tag = article.find("div", class_="date")
            if date_tag:
                date_text = date_tag.text.strip()
                # 将日期与年份结合，生成新的日期格式
                data["日期"] = f"{current_year}/{date_text}"
            else:
                data["日期"] = "N/A"
            data["標題與連結"] = title_link
            data_list.append(data)

    prev_page = soup.find("a", string="‹ 上頁")
    if prev_page and 'href' in prev_page.attrs:
        prev_page_url = "https://www.ptt.cc" + prev_page['href']
    else:
        prev_page_url = None

    return data_list, prev_page_url

def convert_popularity(popularity):
    if popularity == "N/A":
        return 0
    try:
        return int(popularity)
    except ValueError:
        return 0

def save_to_excel(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "PTT Data"
    
    headers = ["人氣", "日期", "標題與連結"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    for row in ws.iter_rows(min_row=2):
        cell = row[2]
        if "(" in cell.value and ")" in cell.value:
            link_start = cell.value.index("(")
            link_end = cell.value.index(")")
            link_url = cell.value[link_start + 1:link_end]
            cell.hyperlink = link_url
            cell.value = cell.value[:link_start].strip()
            cell.font = Font(color="0000FF", underline="single")

    column_widths = []
    for row in dataframe_to_rows(df, index=False, header=True):
        for i, cell in enumerate(row):
            if len(column_widths) < i + 1:
                column_widths.append(len(str(cell)) + 2)
            elif len(str(cell)) + 2 > column_widths[i]:
                column_widths[i] = len(str(cell)) + 2
    
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    wb.save(filename)

class PttCrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("PTT Crawler")

        # GUI
        self.board_label = tk.Label(root, text="請選擇PTT板塊：")
        self.board_label.pack()

        self.board_var = tk.StringVar()
        self.board_menu = ttk.Combobox(root, textvariable=self.board_var)
        self.board_menu['values'] = ("SuperBike", "Gossiping", "Stock", "NBA","Soft_Job","Tech_Job","Beauty")
        self.board_menu.current(0)
        self.board_menu.pack()

        self.max_pages_label = tk.Label(root, text="請輸入要爬取的最大頁數（如果不輸入則爬取所有頁數）：")
        self.max_pages_label.pack()

        self.max_pages_entry = tk.Entry(root)
        self.max_pages_entry.pack()

        self.keyword_label = tk.Label(root, text="請輸入要搜尋的關鍵字（如果不需要過濾請直接按 Enter）：")
        self.keyword_label.pack()

        self.keyword_entry = tk.Entry(root)
        self.keyword_entry.pack()

        self.progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(pady=10)

        self.progress_label = tk.Label(root, text="")
        self.progress_label.pack()

        self.time_remaining_label = tk.Label(root, text="")
        self.time_remaining_label.pack()

        self.start_button = tk.Button(root, text="開始爬取", command=self.start_crawling)
        self.start_button.pack(pady=20)

    def start_crawling(self):
        board = self.board_var.get()
        max_pages = self.max_pages_entry.get()
        keyword = self.keyword_entry.get()

        if max_pages:
            max_pages = int(max_pages)
        else:
            max_pages = -1

        if not keyword:
            keyword = ""

        self.crawl_thread = threading.Thread(target=self.crawl_ptt, args=(board, max_pages, keyword))
        self.crawl_thread.start()

    def crawl_ptt(self, board, max_pages, keyword):
        start_url = f"https://www.ptt.cc/bbs/{board}/index.html"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0'}
        
        session = requests.Session()
        session.post('https://www.ptt.cc/ask/over18', headers=headers, data={'yes': 'yes'})

        if max_pages == -1:
            total_pages = self.get_total_pages(start_url, headers, session)
        else:
            total_pages = max_pages

        all_data = []
        url = start_url
        page_count = 0
        start_time = time.time()

        for i in range(total_pages):
            if not url:
                break

            data_list, url = get_articles_from_page(url, headers, keyword, session)
            all_data.extend(data_list)
            page_count += 1

            # 更新進度條
            progress = int((page_count / total_pages) * 100)
            self.progress['value'] = progress
            self.progress_label['text'] = f"進度: {progress}%"

            elapsed_time = time.time() - start_time
            remaining_time = (elapsed_time / page_count) * (total_pages - page_count)
            self.time_remaining_label['text'] = f"剩餘時間: {str(timedelta(seconds=int(remaining_time)))}"

            self.root.update_idletasks()

        # 爬取完成後更新狀態
        self.progress_label['text'] = "進度: 100% - 爬取完成"
        self.time_remaining_label['text'] = ""

        downloads_path = str(Path.home() / "Downloads")

        if all_data:
            df = pd.DataFrame(all_data)
            if "人氣" in df.columns:
                df["人氣"] = df["人氣"].apply(convert_popularity)
                df = df.sort_values(by="人氣", ascending=False)
                try:
                    excel_path = os.path.join(downloads_path, f"ptt_{board}_sorted.xlsx")
                    save_to_excel(df, excel_path)
                    json_path = os.path.join(downloads_path, f"ptt_{board}_data_sorted.json")
                    with open(json_path, "w", encoding="utf-8") as file:
                        json.dump(all_data, file, ensure_ascii=False, indent=4)
                    messagebox.showinfo("完成", f"資料已經儲存為 {excel_path} 和 {json_path}")
                except PermissionError:
                    messagebox.showerror("錯誤", "無法寫入文件，請檢查文件是否被打開或是否有寫入權限。")
            else:
                messagebox.showinfo("完成", "沒有找到任何符合條件的文章")
        else:
            messagebox.showinfo("完成", "沒有找到任何符合條件的文章")

    def get_total_pages(self, start_url, headers, session):
        response = session.get(start_url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")
        prev_page = soup.find("a", string="‹ 上頁")
        if prev_page and 'href' in prev_page.attrs:
            prev_page_url = prev_page['href']
            total_pages = int(prev_page_url.split("index")[1].split(".html")[0]) + 1
        else:
            total_pages = 1
        return total_pages

def main():
    root = tk.Tk()
    app = PttCrawlerGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
